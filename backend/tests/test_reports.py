"""Tests for the board-pack export — the pure workbook builder and the download endpoint.

The builder is verified by opening the produced bytes with openpyxl and asserting structure/values;
the endpoint is verified end-to-end on isolated accounts so it is independent of other fixtures.
"""

from __future__ import annotations

import io
from decimal import Decimal

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.db.session import init_db
from app.domain.reporting import BoardRow, ReportMeta, build_board_pack
from app.domain.variance import BridgeStep, VarianceItem, build_bridge, build_insights, compose_narrative
from app.main import app
from app.services.variance_query import inr_compact

D = Decimal

init_db()
client = TestClient(app)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _find_row(ws: object, text: str) -> int | None:
    for row in ws.iter_rows(min_col=1, max_col=1):  # type: ignore[attr-defined]
        if row[0].value == text:
            return int(row[0].row)
    return None


def _sample() -> tuple[ReportMeta, list[BoardRow], object, str, object]:
    items = [
        VarianceItem("4000", "Sales", "Revenue", D("20000"), D("120000"), D("100000"), True),
        VarianceItem("6100", "Marketing", "Opex", D("-5000"), D("30000"), D("25000"), False),
    ]
    insight = build_insights(items)
    narrative = compose_narrative(insight, inr_compact)
    bridge = build_bridge(D("100000"), [BridgeStep("Revenue", D("20000")), BridgeStep("Opex", D("-5000"))])
    rows = [
        BoardRow("4000", "Sales", "Revenue", "2027-06", D("100000"), D("120000"), D("20000"), 20.0, "FAVORABLE", True),
        BoardRow("6100", "Marketing", "Opex", "2027-06", D("25000"), D("30000"), D("5000"), 20.0, "UNFAVORABLE", False),
    ]
    meta = ReportMeta(
        title="Variance Analysis — Board Pack",
        subtitle="Budget vs actual with ranked drivers and a contribution bridge",
        base_label="ACTUAL",
        compare_label="BUDGET",
        period_label="2027-06",
        generated_at="2026-06-23 12:00 UTC",
    )
    return meta, rows, insight, narrative, bridge


# --- pure builder ----------------------------------------------------------


def test_build_board_pack_structure() -> None:
    meta, rows, insight, narrative, bridge = _sample()
    content = build_board_pack(meta, rows, insight, narrative, bridge)
    assert content[:2] == b"PK"  # xlsx is a zip archive

    wb = load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["Summary", "Variance detail", "Drivers", "Bridge"]

    summary = wb["Summary"]
    assert summary["A1"].value == "Variance Analysis — Board Pack"

    detail = wb["Variance detail"]
    assert [detail.cell(row=1, column=c).value for c in (1, 5, 6, 7, 9)] == [
        "Account", "Budget", "Actual", "Variance", "Status",
    ]
    # first data row: actual 120000, raw variance 20000, favourable status
    assert detail.cell(row=2, column=6).value == 120000.0
    assert detail.cell(row=2, column=7).value == 20000.0
    assert detail.cell(row=2, column=9).value == "Favourable"

    # totals row sums raw variance: 20000 + 5000 = 25000
    total_row = _find_row(detail, "Total")
    assert total_row is not None
    assert detail.cell(row=total_row, column=7).value == 25000.0


def test_board_pack_bridge_reconciles() -> None:
    meta, rows, insight, narrative, bridge = _sample()
    wb = load_workbook(io.BytesIO(build_board_pack(meta, rows, insight, narrative, bridge)))
    b = wb["Bridge"]
    assert b["A1"].value == "Contribution bridge"
    end_row = _find_row(b, "End — Actual")
    assert end_row is not None
    assert b.cell(row=end_row, column=3).value == 115000.0  # 100000 + 20000 - 5000


def test_board_pack_drivers_sheet_lists_sections() -> None:
    meta, rows, insight, narrative, bridge = _sample()
    wb = load_workbook(io.BytesIO(build_board_pack(meta, rows, insight, narrative, bridge)))
    labels = {row[0].value for row in wb["Drivers"].iter_rows(min_col=1, max_col=1)}
    assert "Top drags (largest unfavourable)" in labels
    assert "Top offsets (largest favourable)" in labels
    assert "By category" in labels


# --- endpoint --------------------------------------------------------------


def _make_account(code: str, name: str, atype: str) -> None:
    client.post("/api/accounts", json={"account_code": code, "account_name": name, "account_type": atype})


def test_variance_pack_endpoint() -> None:
    _make_account("4300", "Subscriptions", "REVENUE")
    _make_account("6400", "Travel", "OPEX")
    period = "2027-09"  # isolated period -> independent of other fixtures
    header = "account_code,entity_code,department_code,project_code,region_code,currency,2027-09\n"
    actual = header + "4300,ALL,UNALLOC,NONE,ALL,INR,140000\n6400,ALL,UNALLOC,NONE,ALL,INR,32000\n"
    budget = header + "4300,ALL,UNALLOC,NONE,ALL,INR,120000\n6400,ALL,UNALLOC,NONE,ALL,INR,30000\n"
    client.post("/api/uploads?scenario=ACTUAL", files={"file": ("a.csv", io.BytesIO(actual.encode()), "text/csv")})
    client.post("/api/uploads?scenario=BUDGET", files={"file": ("b.csv", io.BytesIO(budget.encode()), "text/csv")})

    r = client.post(
        "/api/reports/variance-pack.xlsx",
        json={"base_scenario": "ACTUAL", "compare_scenario": "BUDGET", "period_from": period, "period_to": period},
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX_MIME
    assert "variance-board-pack-actual-vs-budget.xlsx" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Summary", "Variance detail", "Drivers", "Bridge"]
    codes = {row[0].value for row in wb["Variance detail"].iter_rows(min_row=2, min_col=1, max_col=1)}
    assert {"4300", "6400"} <= codes
