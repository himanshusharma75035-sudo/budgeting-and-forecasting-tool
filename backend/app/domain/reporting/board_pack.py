"""Board-ready variance report pack as a formatted Excel workbook (roadmap 0.8).

A *board pack* is the staple FP&A deliverable: a clean, multi-tab ``.xlsx`` finance teams can drop
straight into a board deck. This module is pure — it takes already-computed variance data and emits
workbook ``bytes`` with no database, network, or filesystem access, so it is fully unit-testable.

Tabs:
    Summary         — title block, basis/period meta, headline figures, narrative commentary.
    Variance detail — per-(account, period) budget-vs-actual table with totals and a status column.
    Drivers         — ranked "top drags / top offsets" and a by-category breakdown.
    Bridge          — the category contribution waterfall (start → categories → end).

Money is rendered with Indian digit grouping (``#,##,##0`` = lakh/crore) and the ₹ symbol; the cell
value keeps full precision while the format controls display. Favourable/unfavourable polarity is
shown with colour, never inferred from a raw sign by the reader.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.variance import Bridge, InsightDriver, VarianceInsight

# --- inputs ----------------------------------------------------------------


@dataclass(frozen=True)
class ReportMeta:
    title: str
    subtitle: str
    base_label: str  # e.g. "ACTUAL"
    compare_label: str  # e.g. "BUDGET"
    period_label: str  # e.g. "2026-04 – 2027-03" or "All periods"
    generated_at: str  # caller-supplied display string (keeps the builder pure/deterministic)
    currency: str = "INR"


@dataclass(frozen=True)
class BoardRow:
    account_code: str
    account_name: str
    category: str
    period: str
    comparison: Decimal
    actual: Decimal
    variance: Decimal  # raw: actual - comparison
    variance_pct: float | None  # percentage value, e.g. -7.7 means -7.7%
    status: str  # "FAVORABLE" | "UNFAVORABLE" | "NEUTRAL"
    is_material: bool


# --- styling ---------------------------------------------------------------

_INK = "1F2937"  # gray-800 (header bar)
_WHITE = "FFFFFF"
_MUTED = "6B7280"  # gray-500
_GREEN = "15803D"  # green-700 (favourable)
_RED = "B91C1C"  # red-700 (unfavourable)
_BAND = "F9FAFB"  # gray-50 (zebra band)
_RULE = "E5E7EB"  # gray-200 (borders)
_PANEL = "F3F4F6"  # gray-100 (section labels)

_MONEY_FMT = '"₹"#,##,##0;"₹"-#,##,##0'  # Indian grouping, plain (polarity shown via colour)
_PCT_FMT = "+0.0%;-0.0%;0.0%"

_HEADER_FONT = Font(bold=True, color=_WHITE, size=10)
_HEADER_FILL = PatternFill("solid", fgColor=_INK)
_TITLE_FONT = Font(bold=True, size=18, color=_INK)
_SUBTITLE_FONT = Font(size=11, color=_MUTED)
_LABEL_FONT = Font(bold=True, color=_MUTED, size=10)
_SECTION_FONT = Font(bold=True, size=12, color=_INK)
_SECTION_FILL = PatternFill("solid", fgColor=_PANEL)
_BOLD = Font(bold=True, color=_INK)
_GREEN_FONT = Font(color=_GREEN, bold=True)
_RED_FONT = Font(color=_RED, bold=True)
_BAND_FILL = PatternFill("solid", fgColor=_BAND)
_TOP_RULE = Border(top=Side(style="thin", color=_INK))
_BOTTOM_RULE = Border(bottom=Side(style="thin", color=_RULE))

_RIGHT = Alignment(horizontal="right")
_LEFT = Alignment(horizontal="left")
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_STATUS_LABEL = {"FAVORABLE": "Favourable", "UNFAVORABLE": "Unfavourable", "NEUTRAL": "On plan"}


def _f(value: Decimal) -> float:
    """Decimal → float for openpyxl (display is rounded by the number format; precision is ample)."""
    return float(value)


def _money(ws: Worksheet, row: int, col: int, value: Decimal, *, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=_f(value))
    cell.number_format = _MONEY_FMT
    cell.alignment = _RIGHT
    if bold:
        cell.font = _BOLD


def _signed_money(ws: Worksheet, row: int, col: int, value: Decimal) -> None:
    """Money cell coloured by favourable (≥0 → green) / unfavourable (<0 → red) polarity."""
    cell = ws.cell(row=row, column=col, value=_f(value))
    cell.number_format = _MONEY_FMT
    cell.alignment = _RIGHT
    cell.font = _GREEN_FONT if value >= 0 else _RED_FONT


def _header(ws: Worksheet, row: int, headers: Sequence[str]) -> None:
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _RIGHT if c > 1 else _LEFT


def _section(ws: Worksheet, row: int, text: str, span: int) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = _SECTION_FILL


def _widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# --- sheets ----------------------------------------------------------------


def _summary_sheet(ws: Worksheet, meta: ReportMeta, rows: Sequence[BoardRow],
                   insight: VarianceInsight, narrative: str) -> None:
    _widths(ws, {"A": 26, "B": 22, "C": 16, "D": 16, "E": 16, "F": 16})
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = meta.title
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:F2")
    ws["A2"] = meta.subtitle
    ws["A2"].font = _SUBTITLE_FONT

    meta_rows = [
        ("Basis", f"{meta.base_label} vs {meta.compare_label}"),
        ("Period", meta.period_label),
        ("Currency", meta.currency),
        ("Generated", meta.generated_at),
    ]
    r = 4
    for label, value in meta_rows:
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    _section(ws, r, "Key figures", 6)
    r += 1
    fav_lines = sum(1 for x in rows if x.status == "FAVORABLE")
    unfav_lines = sum(1 for x in rows if x.status == "UNFAVORABLE")
    material_lines = sum(1 for x in rows if x.is_material)
    ws.cell(row=r, column=1, value="Net variance").font = _LABEL_FONT
    _signed_money(ws, r, 2, insight.net_favorable)
    r += 1
    ws.cell(row=r, column=1, value="Favourable total").font = _LABEL_FONT
    _signed_money(ws, r, 2, insight.favorable_total)
    r += 1
    ws.cell(row=r, column=1, value="Unfavourable total").font = _LABEL_FONT
    _signed_money(ws, r, 2, insight.unfavorable_total)
    r += 1
    for label, count in (
        ("Favourable lines", fav_lines),
        ("Unfavourable lines", unfav_lines),
        ("Material lines", material_lines),
    ):
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=2, value=count).alignment = _RIGHT
        r += 1

    r += 1
    _section(ws, r, "Commentary", 6)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 5, end_column=6)
    cell = ws.cell(row=r, column=1, value=narrative.replace("**", ""))
    cell.alignment = _WRAP


def _detail_sheet(ws: Worksheet, meta: ReportMeta, rows: Sequence[BoardRow]) -> None:
    headers = [
        "Account", "Name", "Category", "Period",
        meta.compare_label.title(), meta.base_label.title(), "Variance", "Var %", "Status",
    ]
    _header(ws, 1, headers)
    _widths(ws, {"A": 12, "B": 26, "C": 18, "D": 10, "E": 16, "F": 16, "G": 16, "H": 10, "I": 14})
    ws.freeze_panes = "A2"

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row.account_code)
        ws.cell(row=r, column=2, value=row.account_name)
        ws.cell(row=r, column=3, value=row.category or "—")
        ws.cell(row=r, column=4, value=row.period)
        _money(ws, r, 5, row.comparison)
        _money(ws, r, 6, row.actual)
        _money(ws, r, 7, row.variance)
        if row.variance_pct is not None:
            pct = ws.cell(row=r, column=8, value=row.variance_pct / 100.0)
            pct.number_format = _PCT_FMT
            pct.alignment = _RIGHT
        status = ws.cell(row=r, column=9, value=_STATUS_LABEL.get(row.status, row.status.title()))
        if row.status == "FAVORABLE":
            status.font = _GREEN_FONT
        elif row.status == "UNFAVORABLE":
            status.font = _RED_FONT
        if r % 2 == 1:  # zebra banding
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = _BAND_FILL
        r += 1

    # totals row
    total_comparison = sum((x.comparison for x in rows), Decimal("0"))
    total_actual = sum((x.actual for x in rows), Decimal("0"))
    total_variance = sum((x.variance for x in rows), Decimal("0"))
    label = ws.cell(row=r, column=1, value="Total")
    label.font = _BOLD
    _money(ws, r, 5, total_comparison, bold=True)
    _money(ws, r, 6, total_actual, bold=True)
    _money(ws, r, 7, total_variance, bold=True)
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = _TOP_RULE

    ws.auto_filter.ref = f"A1:I{max(1, len(rows) + 1)}"


def _driver_table(ws: Worksheet, start_row: int, title: str,
                  drivers: Sequence[InsightDriver]) -> int:
    """Render a 'label | category | impact' driver table; returns the next free row."""
    _section(ws, start_row, title, 3)
    _header(ws, start_row + 1, ["Account", "Category", "Impact"])
    r = start_row + 2
    for d in drivers:
        ws.cell(row=r, column=1, value=d.label)
        ws.cell(row=r, column=2, value=d.category or "—")
        _signed_money(ws, r, 3, d.favorable_variance)
        r += 1
    if not drivers:
        ws.cell(row=r, column=1, value="—").font = Font(color=_MUTED)
        r += 1
    return r + 1


def _drivers_sheet(ws: Worksheet, insight: VarianceInsight) -> None:
    _widths(ws, {"A": 28, "B": 22, "C": 18})
    ws.sheet_view.showGridLines = False
    r = 1
    r = _driver_table(ws, r, "Top drags (largest unfavourable)", insight.top_unfavorable)
    r = _driver_table(ws, r, "Top offsets (largest favourable)", insight.top_favorable)

    _section(ws, r, "By category", 4)
    _header(ws, r + 1, ["Category", "Net impact", "Actual", "Budget"])
    ws.column_dimensions["D"].width = 16
    rr = r + 2
    for d in insight.by_category:
        ws.cell(row=rr, column=1, value=d.label)
        _signed_money(ws, rr, 2, d.favorable_variance)
        _money(ws, rr, 3, d.actual)
        _money(ws, rr, 4, d.comparison)
        rr += 1


def _bridge_sheet(ws: Worksheet, meta: ReportMeta, bridge: Bridge) -> None:
    _widths(ws, {"A": 32, "B": 18, "C": 18})
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Contribution bridge"
    ws["A1"].font = _SECTION_FONT
    _header(ws, 2, ["Step", "Contribution", "Running total"])

    r = 3
    running = bridge.start
    start_cell = ws.cell(row=r, column=1, value=f"Start — {meta.compare_label.title()}")
    start_cell.font = _BOLD
    _money(ws, r, 3, running, bold=True)
    r += 1
    for step in bridge.steps:
        ws.cell(row=r, column=1, value=step.label)
        _signed_money(ws, r, 2, step.delta)
        running += step.delta
        _money(ws, r, 3, running)
        r += 1
    end_cell = ws.cell(row=r, column=1, value=f"End — {meta.base_label.title()}")
    end_cell.font = _BOLD
    _money(ws, r, 3, bridge.end, bold=True)
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = _TOP_RULE


# --- entry point -----------------------------------------------------------


def build_board_pack(
    meta: ReportMeta,
    rows: Sequence[BoardRow],
    insight: VarianceInsight,
    narrative: str,
    bridge: Bridge,
) -> bytes:
    """Assemble the four-tab board pack and return the ``.xlsx`` file as bytes."""
    wb = Workbook()
    wb.properties.creator = "OpenFPA"
    wb.properties.title = meta.title

    summary = wb.active
    assert isinstance(summary, Worksheet)  # a fresh Workbook always has exactly one worksheet
    summary.title = "Summary"
    _summary_sheet(summary, meta, rows, insight, narrative)
    _detail_sheet(wb.create_sheet("Variance detail"), meta, rows)
    _drivers_sheet(wb.create_sheet("Drivers"), insight)
    _bridge_sheet(wb.create_sheet("Bridge"), meta, bridge)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
